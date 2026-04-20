"""
Re-parser for Semrush organic positions that also captures the "Position Type"
(Organic vs SERP feature type like People also ask, Image pack, Sitelinks, etc.).

Logic:
 - Each snapshot YAML file contains two representations of the table:
   (1) an "interactive" section with links/buttons per row, including an optional
       link named "Current position: <feature>" when the row is ranked AS a SERP feature
   (2) a "gridcells" section with plain text data per cell (keyword, intent, pos, SF,
       traffic, traffic%, volume, KD, URL)
 - We use (1) to get the feature type for each keyword (in order) and (2) to get
   the numeric data. Joining is done by row index (both sections list the same
   keywords in the same order).
"""

import re
import openpyxl
from pathlib import Path

SNAPSHOTS = [
    r"C:\Users\Usuario\.cursor\browser-logs\snapshot-2026-04-20T12-30-12-807Z-za7fub.log",  # page 1
    r"C:\Users\Usuario\.cursor\browser-logs\snapshot-2026-04-20T12-30-40-378Z-1i1x2x.log",  # page 2
    r"C:\Users\Usuario\.cursor\browser-logs\snapshot-2026-04-20T12-31-22-679Z-19617t.log",  # page 3
]

OUTPUT = r"C:\Users\Usuario\OneDrive\Documentos\seo 2026\repediu_semrush_positions_202503_v2.xlsx"


RE_ANALYZE_KW = re.compile(
    r"- role: link\s*\n\s+name: Analyze (.+?) in Keyword Overview",
    re.MULTILINE,
)
RE_CURRENT_POS = re.compile(
    r'- role: link\s*\n\s+name: "Current position: (.+?)"',
    re.MULTILINE,
)
RE_GRIDCELL = re.compile(
    r"- role: gridcell\s*\n\s+name: (.+?)\n\s+ref: e\d+",
    re.MULTILINE,
)


def extract_feature_types(text: str):
    """Extract an ordered list of (keyword, feature_type_or_None) tuples.

    We walk the text in order and find all "Analyze X in Keyword Overview" links
    (which mark the start of each row in the interactive section). For each
    keyword, we look in the slice of text until the NEXT "Analyze ... in Keyword
    Overview" link and check whether a "Current position: XXX" link exists.

    Note: The text also contains trailing "Analyze <url>" links for the URL
    column. We must only match keywords, so we filter using the " in Keyword
    Overview" suffix.
    """
    kw_matches = list(RE_ANALYZE_KW.finditer(text))
    result = []
    for i, m in enumerate(kw_matches):
        kw = m.group(1).strip()
        start = m.end()
        end = kw_matches[i + 1].start() if i + 1 < len(kw_matches) else len(text)
        segment = text[start:end]
        fm = RE_CURRENT_POS.search(segment)
        feature = fm.group(1).strip() if fm else None
        result.append((kw, feature))
    return result


def extract_gridcells(text: str):
    cells = []
    for m in RE_GRIDCELL.finditer(text):
        val = m.group(1).strip()
        if val.startswith('"') and val.endswith('"'):
            val = val[1:-1]
        cells.append(val)
    return cells


def parse_rows_from_cells(cells):
    """Group flat gridcell values into rows of 8 or 9 cells.

    A row ends with a URL cell that starts with "repediu.com.br".
    """
    rows = []
    current = []
    for c in cells:
        current.append(c)
        if c.startswith("repediu.com.br") or c == "repediu.com.br/":
            rows.append(current)
            current = []
    return rows


def normalize_row(row):
    """Ensure each row has 9 columns: KW, Intent, Pos, SF, Traffic, Traffic%, Volume, KD, URL.

    Rows that are ranked as a SERP feature often DON'T have an SF count cell
    (because Semrush renders an icon there). We pad SF with "" in that case.
    """
    if len(row) == 9:
        return row
    if len(row) == 8:
        # Insert empty SF at index 3
        return row[:3] + [""] + row[3:]
    if len(row) < 9:
        return row[:-1] + [""] * (9 - len(row)) + [row[-1]]
    return row[:9]


def main():
    all_enriched_rows = []

    for page_idx, snap_path in enumerate(SNAPSHOTS, 1):
        text = Path(snap_path).read_text(encoding="utf-8")

        kw_features = extract_feature_types(text)
        cells = extract_gridcells(text)
        rows = parse_rows_from_cells(cells)
        rows = [normalize_row(r) for r in rows if r and r[-1].startswith("repediu.com.br")]

        print(f"Page {page_idx}: {len(rows)} rows | {len(kw_features)} kw-feature tags")

        if len(rows) != len(kw_features):
            print(
                f"  WARN: row/tag count mismatch on page {page_idx} "
                f"({len(rows)} rows vs {len(kw_features)} feature tags)"
            )

        for i, row in enumerate(rows):
            kw_from_grid = row[0].strip().lower()
            if i < len(kw_features):
                kw_from_int, feature = kw_features[i]
                if kw_from_int.strip().lower() != kw_from_grid:
                    print(
                        f"  WARN: keyword mismatch at row {i}: "
                        f"interactive='{kw_from_int}' vs grid='{row[0]}'"
                    )
            else:
                feature = None
            position_type = feature if feature else "Organic"
            has_icon = "Sim" if feature else "Nao"
            all_enriched_rows.append(row + [position_type, has_icon])

    print(f"Total rows: {len(all_enriched_rows)}")

    feature_counts = {}
    for r in all_enriched_rows:
        t = r[-2]
        feature_counts[t] = feature_counts.get(t, 0) + 1
    print("Breakdown by Position Type:")
    for t, n in sorted(feature_counts.items(), key=lambda x: -x[1]):
        print(f"  {t}: {n}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Positions March 2025"
    headers = [
        "Keyword",
        "Intent",
        "Position",
        "SF (count)",
        "Traffic",
        "Traffic %",
        "Volume",
        "KD %",
        "URL",
        "Position Type",
        "Tem icone?",
    ]
    ws.append(headers)
    for r in all_enriched_rows:
        ws.append(r)

    # Auto-size columns
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = min(max_len + 2, 60)

    # Freeze header
    ws.freeze_panes = "A2"

    # Enable auto-filter on header row
    ws.auto_filter.ref = ws.dimensions

    # Additional sheet: summary per Position Type
    ws2 = wb.create_sheet("Resumo Tipo Posicao")
    ws2.append(["Position Type", "Qtd keywords", "% do total"])
    total = len(all_enriched_rows)
    for t, n in sorted(feature_counts.items(), key=lambda x: -x[1]):
        pct = (n / total * 100) if total else 0
        ws2.append([t, n, f"{pct:.2f}%"])
    for col_idx, col_cells in enumerate(ws2.columns, 1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws2.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max_len + 2

    Path(OUTPUT).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
